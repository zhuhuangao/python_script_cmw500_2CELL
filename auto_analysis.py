from PyPDF2 import PdfMerger
import os
import re
import math
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from pandas.io import json
from reportlab.pdfgen import canvas
import matplotlib.cm as cm
from openpyxl.styles import Border, Side
import matplotlib.backends.backend_pdf
from collections import Counter
from collections import OrderedDict

plt.rcParams["font.sans-serif"] = ["SimHei"]
plt.rcParams["axes.unicode_minus"] = False

# 1. 获取当前目录文件名包含“Test_PSM”的.csv文件
file_list = [file for file in os.listdir('.') if 'Test' in file and file.endswith('.csv')]
if len(file_list) == 0:
    print("未找到符合条件的文件。")
    exit()
#  获取当前目录文件名包含“Test_PSM”的.csv文件
file_png = [file for file in os.listdir('.') if 'PSM' in file and file.endswith('.png')]
file_list2 = [file for file in os.listdir('.') if 'Pcorelog' in file and file.endswith('.txt')]

file_list3 = [file for file in os.listdir('.') if 'my_report' in file and file.endswith('.txt')]


# 获取Excel里面的数据分析
def psm_analys():
    global version, data_dict, dict2, data_dict0, data_dict3
    # 2. 读取名字为“PSM模式测试”的sheet
    df = pd.read_excel(file_list[0], sheet_name='PSM模式测试', header=None)
    df1 = pd.read_excel(file_list[0], sheet_name='测试数据', header=None)
    version1 = df1.iat[0, 0]
    version2 = df1.iat[0, 1]
    data_dict[version1] = version2

    # 3. 获取第4列中除第1,2行以外的所有数据，写入到数组list1中，并将其转化成float类型，获取PSM入网时间
    list1 = df.iloc[2:, 3].astype(float).tolist()
    list1_key = df.iloc[1, 3]
    dict2 = {list1_key: list1}

    # 4. 获取第5列中除第1,2行以外的所有数据,频点，写入到数组list2中，获取频点
    list2 = df.iloc[2:, 4].tolist()
    list2_key = df.iloc[1, 4]
    dict2.update({list2_key: list2})

    # 获取第6列中除第1,2行以外的所有数据，写入到数组list3中，获取小区
    list3 = df.iloc[2:, 5].tolist()

    # RSRP
    dict2.update({df.iloc[1, 7]: df.iloc[2:, 7].astype(float).tolist()})
    # SNR
    dict2.update({df.iloc[1, 8]: df.iloc[2:, 8].astype(float).tolist()})

    # 获取第11列中除第1,2行以外的所有数据，写入到数组list4中，获取小区，云平台注册时间（包括登录，订阅）
    list4 = df.iloc[2:, 10].astype(float).tolist()

    # 获取第12列中除第1,2行以外的所有数据，写入到数组list5中，获取小区，云平台发送数据时间
    list5 = df.iloc[2:, 11].astype(float).tolist()
    # 5. 对list1进行统计并生成饼图----------------------------------------------------------------------
    data_range = [-1, 0, 5, 10, float('inf')]
    pie_data1 = [sum(1 for num in list1 if data_range[i] <= num < data_range[i+1]) for i in range(len(data_range)-1)]
    labels1 = ['不入网', '0-5s', '5-10s', '10s+']
    plt.subplot(231)
    colors1 = ['lightblue', 'lightgreen', 'lightyellow', 'lightcoral']
    plt.pie(pie_data1, explode=[0.01, 0.01, 0.01, 0.01], labels=labels1, labeldistance=1.1, autopct='%1.1f%%')
    plt.title(f'PSM入网时间分布({len(list1)})')
    plt.legend(labels1, loc='upper right', fontsize=5)

    data_dict3.update({'PSM入网时间分布': None})
    for j in range(len(labels1)):
        data_rate = format(float(pie_data1[j]/len(list1))*100,'.2f')
        pie_data1[j] = str(pie_data1[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels1)):
        data_dict3[labels1[j]] = pie_data1[j]

    # 6. 对list2进行统计并生成饼图-------------------------------------------------------
    counts = {}
    list_t = []
    for i in list2:
        list_t.append(str(i))
    list2 = list_t
    for item in list2:
        if item in counts:
            counts[item] += 1
        else:
            counts[item] = 1

    labels2 = list(counts.keys())
    # print(labels2)
    values2 = list(counts.values())
    # print(counts)
    explode = []
    for i in range(len(counts)):
        explode.append(0.01)

    plt.subplot(232)
    plt.pie(values2, explode=explode, labels=labels2, autopct='%1.1f%%')
    plt.title(f'入网频点分布({len(list2)})')
    plt.legend(labels2, loc='upper right', fontsize=5)

    data_dict3.update({'入网频点分布': None})

    for j in range(len(labels2)):
        data_rate = format(float(values2[j]/len(list2))*100,'.2f')
        values2[j] = str(values2[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels2)):
        data_dict3[labels2[j]] = values2[j]

    # 统计list3数据，生成饼图----------------------------------------------------------------
    counts2 = {}
    for item in list3:
        if item in counts2:
            counts2[item] += 1
        else:
            counts2[item] = 1

    labels3 = list(counts2.keys())
    values3 = list(counts2.values())
    explode = []
    for i in range(len(counts2)):
        explode.append(0.01)

    plt.subplot(233)
    plt.pie(values3, explode=explode, labels=labels3, autopct='%1.1f%%')
    plt.title(f'入网小区分布({len(list3)})次')
    plt.legend(labels3, loc='upper right', fontsize=5)

    data_dict3.update({'入网小区分布': None})

    for j in range(len(labels3)):
        data_rate = format(float(values3[j] / len(list3))*100, '.2f')
        values3[j] = str(values3[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels2)):
        data_dict3[labels3[j]] = values3[j]

    # # 获取第11列中除第1,2行以外的所有数据，写入到数组list4中，云平台注册时间（包括登录，订阅），生成饼图--------------------------------------
    data_range2 = [-1, 0.1, 3, float('inf')]
    pie_data1 = [sum(1 for num in list4 if data_range2[i] <= num < data_range2[i+1]) for i in range(len(data_range2)-1)]
    labels1 = ['-1-0', '0-3s', '3s+']

    plt.subplot(234)
    colors1 = ['lightblue', 'lightgreen', 'lightyellow', 'lightcoral']
    plt.pie(pie_data1, explode=[0.01, 0.01, 0.01], labels=labels1, labeldistance=1.1, autopct='%1.1f%%')

    plt.title(f'注册登录时间分布({len(list4)})次')
    plt.legend(labels1, loc='upper right', fontsize=5)

    data_dict3.update({'注册登录时间分布':  None})
    for j in range(len(labels1)):
        data_rate = format(float(pie_data1[j]/len(list1))*100,'.2f')
        pie_data1[j] = str(pie_data1[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels1)):
        data_dict3[labels1[j]] = pie_data1[j]

    # 获取第12列中除第1,2行以外的所有数据，写入到数组list5中，发送数据时间，生成饼图------------------------------------------------------
    data_range5 = [-1, 0.1, 2, float('inf')]
    pie_data5 = [sum(1 for num in list5 if data_range5[i] <= num < data_range5[i+1]) for i in range(len(data_range5)-1)]
    labels5 = ['-1-0s', '0-2s', '2s+']
    # print(pie_data5_str)
    # print(type(pie_data5_str[0]))

    plt.subplot(235)
    # colors1 = ['lightblue', 'lightgreen', 'lightyellow', 'lightcoral']
    plt.pie(pie_data5, explode=[0.01, 0.01, 0.01], labels=labels5, labeldistance=1.1, autopct='%1.1f%%')

    plt.title(f'发送数据时间分布({len(list5)})次')
    plt.legend(labels5, loc='upper right', fontsize=5)

    data_dict3.update({'发送数据时间分布': None})
    for j in range(len(labels5)):
        data_rate = format(float(pie_data5[j] / len(list5))*100, '.2f')
        pie_data5[j] = str(pie_data5[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels5)):
        data_dict3[labels5[j]] = pie_data5[j]


    plt.tight_layout()
    plt.savefig('PSM_data_analist.png')

    # SNR
    data_SNR = pd.Series(dict2['SNR'])
    data_SNR = data_SNR[(data_SNR != 255) & (data_SNR != 25)]
    # data_SNR = data_SNR[(data_SNR != -1) & (data_SNR != 0)]
    # # 计算等于 -1 的元素个数
    # count = data1.value_counts().to_string(index=False, dtype=False)
    # 求平均值
    mean_sub_value = round(data_SNR.mean(), 2)
    # 求最大值
    max_sub_value = round(data_SNR.max(), 2)
    # 求最小值
    min_sub_value = round(data_SNR.min(), 2)
    # 求方差
    var_sub_value = round(data_SNR.var(), 2)
    # 写入字典
    data_dict0["SNR平均值"] = mean_sub_value
    data_dict0["SNR最大值"] = max_sub_value
    data_dict0["SNR最小值"] = min_sub_value
    data_dict0["SNR方差"] = var_sub_value

    # data_sub = pd.Series(doc1['Sub_time'])
    data_total_analysis(data_dic_key='RSRP')
    data_total_analysis(data_dic_key='PSM入网时间')


# 获取excel里面的数据，输出最大，最小、平均、方差
def data_total_analysis(data_dic_key):
    global doc1, data_dict, dict2, data_dict0
    if psm_mode != "1":
        data = pd.Series(doc1[data_dic_key])
    else:
        data = pd.Series(dict2[data_dic_key])
    data = data[(data != -1) & (data != 0)]
    # 求平均值
    mean_value = str(round(data.mean(), 2))
    # 求最大值
    max_value = str(round(data.max(), 2))
    # 求最小值
    min_value = str(round(data.min(), 2))
    # 求方差
    var_value = str(round(data.var(), 2))
    # 写入字典
    data_dict0[data_dic_key + "平均值"] = mean_value
    data_dict0[data_dic_key + "最大值"] = max_value
    data_dict0[data_dic_key + "最小值"] = min_value
    data_dict0[data_dic_key + "方差"] = var_value


# 获取Excel里面的数据分析
def non_psm_analys():
    global data_dict, doc1, data_dict0, dict_sheet2, data_dict3

    # 获取当前目录中所有.csv文件名包含"Test"的文件
    file_list = [f for f in os.listdir('.') if f.endswith('.csv') and 'Test' in f]

    # 遍历所有文件
    for file_name in file_list:
        # 读取名为"测试数据"的sheet
        df = pd.read_excel(file_name, sheet_name='测试数据')

    # 获取第2列中除第1行以外的所有数据, 入网时间写入字典
    key1 = df.iloc[0, 1]
    data1 = df.iloc[1:, 1].astype(float).tolist()
    # 构建字典doc1
    doc1 = {key1: data1}
    # 订阅和注册时间写入字典
    key2 = df.iloc[0, 4]
    data2 = df.iloc[1:, 4].astype(float).tolist()
    doc1.update({key2: data2})
    # 发送数据时间写入字典
    key5 = df.iloc[0, 5]
    data5 = df.iloc[1:, 5].astype(float).tolist()
    doc1.update({key5: data5})

    # 频点写入字典
    key7 = df.iloc[0, 7]
    data7 = df.iloc[1:, 7].tolist()
    doc1.update({key7: data7})

    # 小区写入字典
    key8 = df.iloc[0, 8]
    data8 = df.iloc[1:, 8].tolist()
    doc1.update({key8: data8})

    # RSRP写入字典
    key10 = df.iloc[0, 10]
    data10 = df.iloc[1:, 10].tolist()
    data10 = [float(x) for x in data10]
    doc1.update({key10: data10})

    # SNR写入字典
    key11 = df.iloc[0, 11]
    data11 = df.iloc[1:, 11].tolist()
    data11 = [float(x) for x in data11]
    doc1.update({key11: data11})

    # 丢包率写入字典
    key18 = df.iloc[0, 18]
    data18 = df.iloc[1:, 18].astype(float).tolist()
    doc1.update({key18: data18})
    cell_pingloss_dict = {}
    cell_pingloss_all = {'ping无响应次数': 0}
    for key, value in zip(doc1['CELLID'], doc1['ping_loss']):
        if key in cell_pingloss_dict:
            # if key != '0':
            if value != -1:
                cell_pingloss_dict[key] += value  # 如果键已经存在于 result_dict 中，则将 value 累加到对应的值上
            else:
                cell_pingloss_all['ping无响应次数'] += 1
        else:
            if value != 0 and value != -1:
                # print(cell_pingloss_dict[key])
                cell_pingloss_dict[key] = value  # 如果键在 result_dict 中不存在，则直接添加键值对
                # print(value)
            if value == -1:
                cell_pingloss_all['ping无响应次数'] += 1
    dic_cell_total = dict(Counter(doc1['CELLID']))
    # print(cell_pingloss_dict)
    data_dict.update(cell_pingloss_all)
    for key, value in cell_pingloss_dict.items():
        cell_ping_loss_rate = format(float(int(value)/int(dic_cell_total[key]))*10, '.2f')
        data_dict0[key + '小区丢包率: '] = str(cell_ping_loss_rate) + '%'
        print(key + '小区丢包率: ' + str(cell_ping_loss_rate) + '%')
        # print(value)
        # print(dic_cell_total[key])
    keys = cell_pingloss_dict.keys()  # 获取键
    values = cell_pingloss_dict.values()  # 获取值
    # print(cell_pingloss_dict)

    # print(doc1['CELLID'])
    cell_id_new = '0'
    rsrp_change = 0
    for value in range(len(doc1['CELLID'])):
        cell_id_old = doc1['CELLID'][value]
        if cell_id_new != '0' and cell_id_new != cell_id_old:
            com_result = doc1['RSRP'][value-1] - doc1['RSRP'][value]
            if com_result > 6:
                rsrp_change += 1
        cell_id_new = cell_id_old
    data_dict0['小区变化RSRP变化-6dbm'] = rsrp_change
    colors = cm.rainbow(np.linspace(0, 1, len(keys)))   # 自定义颜色列表，与 keys 一一对应
    # 小区丢包
    # plt.subplot(337)
    # 定义坐标轴
    ax7 = plt1.add_subplot(337)

    # 添加标签
    for key, value in cell_pingloss_dict.items():
        ax7.text(key, value, str(value), ha='center', va='bottom')

    ax7.bar(keys, values, color=colors)  # 绘制柱状图
    ax7.set_xlabel('CELLID')  # 设置 x 轴标签
    ax7.set_ylabel('丢包数')  # 设置 y 轴标签
    ax7.set_title('每个小区丢包数')  # 设置标题

    # 入网时间进行统计并生成饼图----------------------------------------------------------------------
    data_range = [-1, 0.1, 15, 30, float('inf')]
    pie_data1 = [sum(1 for num in doc1["Time"] if data_range[i] <= num < data_range[i + 1]) for i in
                 range(len(data_range) - 1)]
    labels1 = ['不入网', '0-15s', '15-30s', '30s+']

    ax1 = plt1.add_subplot(331)
    # plt.subplot(331)
    colors1 = ['lightblue', 'lightgreen', 'lightyellow', 'lightcoral']
    ax1.pie(pie_data1, explode=[0.02, 0.02, 0.02, 0.02], labels=labels1, labeldistance=1.1, autopct='%1.1f%%')
    ax1.set_title(f'入网时间分布({len(doc1["Time"])})')
    ax1.legend(labels1, loc='upper right', fontsize=3)

    data_dict3.update({'入网时间分布': None})
    for j in range(len(labels1)):
        data_rate = format(float(pie_data1[j]/len(doc1["Time"]))*100,'.2f')
        pie_data1[j] = str(pie_data1[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels1)):
        data_dict3[labels1[j]] = pie_data1[j]

    # 注册订阅时间进行统计并生成饼图----------------------------------------------------------------------
    data_range = [-1, 0.1, 2, 5, float('inf')]
    pie_data2 = [sum(1 for num in doc1["Sub_time"] if data_range[i] <= num < data_range[i + 1]) for i in
                 range(len(data_range) - 1)]
    labels2 = ['订阅失败', '0-2s', '2-5s', '5s+']
    ax2 = plt1.add_subplot(332)
    # plt.subplot(332)
    colors1 = ['lightblue', 'lightgreen', 'lightyellow', 'lightcoral']
    ax2.pie(pie_data2, explode=[0.02, 0.02, 0.02, 0.02], labels=labels2, labeldistance=1.1, autopct='%1.1f%%')
    ax2.set_title(f'订阅注册时间分布({len(doc1["Sub_time"])})')
    ax2.legend(labels1, loc='upper right', fontsize=3)
    data_dict3.update({'订阅注册时间分布': None})
    for j in range(len(labels2)):
        data_rate = float(format(pie_data2[j]/len(doc1["Sub_time"]), '.2f'))*100
        pie_data2[j] = str(pie_data2[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels2)):
        data_dict3[labels2[j]] = pie_data2[j]

    # 云平台发送数据时间进行统计并生成饼图----------------------------------------------------------------------
    data_range = [-1, 0.1, 2, 5, float('inf')]
    pie_data3 = [sum(1 for num in doc1["Send_data_time"] if data_range[i] <= num < data_range[i + 1]) for i in
                 range(len(data_range) - 1)]
    labels3 = ['-1-0', '0s-2s', '2s-5s', '5s+.']
    ax3 = plt1.add_subplot(333)
    # plt.subplot(333)
    colors1 = ['lightblue', 'lightgreen', 'lightyellow', 'lightcoral']
    ax3.pie(pie_data3, explode=[0.02, 0.02, 0.02, 0.02], labels=labels3, labeldistance=1.1, autopct='%1.1f%%')
    ax3.set_title(f'发送数据时间分布({len(doc1["Sub_time"])})')
    ax3.legend(labels3, loc='upper right', fontsize=3)

    data_dict3.update({'发送数据时间分布': None})
    for j in range(len(labels3)):
        data_rate = format(float(pie_data3[j]/len(doc1["Sub_time"]))*100,'.2f')
        pie_data3[j] = str(pie_data3[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels3)):
        data_dict3[labels3[j]] = pie_data3[j]

    # 统计频点分布数据，生成饼图----------------------------------------------------------------
    counts4 = {}
    for item in doc1['频点']:
        if item in counts4:
            counts4[item] += 1
        else:
            counts4[item] = 1

    labels4 = list(counts4.keys())
    values4 = list(counts4.values())
    explode4 = []
    for i in range(len(counts4)):
        explode4.append(0.02)
    ax4 = plt1.add_subplot(334)
    # plt.subplot(334)
    ax4.pie(values4, explode=explode4, labels=labels4, autopct='%1.1f%%')
    ax4.set_title(f'入网频点分布({len(doc1["频点"])})次')
    # plt.legend(labels4, loc='upper right', fontsize=3)

    data_dict3.update({'入网频点分布': None})
    for j in range(len(labels4)):
        data_rate = format(float(values4[j] / len(doc1["频点"]))*100, '.2f')
        values4[j] = str(values4[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels4)):
        data_dict3[labels4[j]] = values4[j]

    # 统计小区分布数据，生成饼图----------------------------------------------------------------
    counts5 = {}
    for item in doc1['CELLID']:
        if item in counts5:
            counts5[item] += 1
        else:
            counts5[item] = 1

    labels5 = list(counts5.keys())
    values5 = list(counts5.values())
    explode5 = []
    for i in range(len(counts5)):
        explode5.append(0.02)
    ax5 = plt1.add_subplot(335)
    # plt.subplot(335)
    ax5.pie(values5, explode=explode5, labels=labels5, autopct='%1.1f%%')
    ax5.set_title(f'入网小区分布({len(doc1["频点"])})次')
    # plt.legend(labels5, loc='upper right', fontsize=3)

    data_dict3.update({'入网小区分布': None})
    for j in range(len(labels5)):
        data_rate = format(float(values5[j] / len(doc1["频点"]))*100, '.2f')
        values5[j] = str(values5[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels5)):
        data_dict3[labels5[j]] = values5[j]

    # 丢包数画成折线图
    ax8 = plt1.add_subplot(338)
    # plt.subplot(338)
    xdata = np.arange(len(doc1['ping_loss']))
    ax8.set_yticks([0, 2, 4, 6, 8, 10])
    ax8.plot(xdata, doc1['ping_loss'], label="Ping_loss", color='#3cb44b', linewidth=0.5)
    # plt.yticks([0, 2, 4, 6, 8, 10])

    dict_cellid_rsrp_snr = {}
    for i in range(len(doc1['CELLID'])):
        listn = [doc1["RSRP"][i], doc1['SNR'][i]]
        if doc1['CELLID'][i] in dict_cellid_rsrp_snr:
            dict_cellid_rsrp_snr[doc1['CELLID'][i]].append(listn )
        else:
            dict_cellid_rsrp_snr[doc1['CELLID'][i]] = [listn]

    # print(len(dict_cellid_rsrp_snr))
    for item in dict_cellid_rsrp_snr:
        plt2 = plt.figure()
        # print(item)
        if item != '0':
            rsrp_i = []
            snr_i = []
            # print(dict_cellid_rsrp_snr[item])
            i = 0
            list_x = []
            for list_i in dict_cellid_rsrp_snr[item]:
                i += 1
                list_x.append(i)
                # print(list_i)
                rsrp_i.append(list_i[0])
                snr_i.append(list_i[1])
            # print(rsrp_i)
            # print(snr_i)
            # print(len(rsrp_i))
            # print(len(snr_i))
            # print(list_x)
            ax_i = plt2.add_subplot()
            ax_i.plot(list_x, rsrp_i, label="RSRP", color='#e6194B', linewidth=0.8)
            ax_i.set_title('CELLID' + str(item))
            ax_ii = ax_i.twinx()
            ax_ii.plot(list_x, snr_i, label="SNR", color='#4363d8', linewidth=0.8)
            ax_i.set_yticks([-60, -80, -90, -100, -110, -120, -130, -140])
            ax_ii.set_yticks([-10, -4, -2, 0, 2, 4, 6, 8, 10, 16, 20, 30])
            ax_i.set_ylim(-140, -60)
            ax_ii.set_ylim(-10, 30)
            ax_ii.legend(loc='upper right', fontsize=6)
            ax_i.legend(loc='upper left', fontsize=6)
            plt2.savefig(str(item) + '小区RSRP-SNR折线图')
            # color = ['#e6194B', '#3cb44b', '#ffe119', '#4363d8']
            # plt1.plot(list_x, snr_i)
            # plt1.savefig(str(item), dpi=300, bbox_inches='tight')

    # print(dict_cellid_rsrp_snr)
    # print(dict_cellid_rsrp_snr['1'][0])


    # print(dict_new)

    ping_result()
    # plt.legend(loc='upper right', fontsize=3)

    plt1.tight_layout()
    # print(type(doc1['Time']))
    plt1.savefig('data_analist.png')

    # print(doc1.keys())
    # print(doc1['ping_loss'])
    i = 0
    add_times = []
    add_loss = []
    add_fre = []
    add_cellid = []
    add_rsrp = []
    add_snr = []
    # print(doc1['ping_loss'])
    for loss in doc1['ping_loss']:
        if loss != 0:
            add_loss.append(loss)
            add_fre.append(doc1['频点'][i])
            add_cellid.append(doc1['CELLID'][i])
            add_rsrp.append(doc1['RSRP'][i])
            add_snr.append(doc1['SNR'][i])
            add_times.append(i+1)
        i += 1
    dict_sheet2['次数'] = add_times
    dict_sheet2['ping_loss'] = add_loss
    dict_sheet2['频点'] = add_fre
    dict_sheet2['CELLID'] = add_cellid
    dict_sheet2['RSRP'] = add_rsrp
    dict_sheet2['SNR'] = add_snr
    # print(dict_sheet2)
    # SNR
    data_SNR = pd.Series(doc1['SNR'])
    data_SNR = data_SNR[(data_SNR != 255) & (data_SNR != 25)]
    # data_SNR = data_SNR[(data_SNR != -1) & (data_SNR != 0)]
    # # 计算等于 -1 的元素个数
    # count = data1.value_counts().to_string(index=False, dtype=False)
    # 求平均值
    mean_sub_value = str(round(data_SNR.mean(), 2))
    # 求最大值
    max_sub_value = str(round(data_SNR.max(), 2))
    # 求最小值
    min_sub_value = str(round(data_SNR.min(), 2))
    # 求方差
    var_sub_value = str(round(data_SNR.var(), 2))
    # 写入字典
    data_dict0["SNR平均值"] = mean_sub_value
    data_dict0["SNR最大值"] = max_sub_value
    data_dict0["SNR最小值"] = min_sub_value
    data_dict0["SNR方差"] = var_sub_value

    # data_sub = pd.Series(doc1['Sub_time'])
    data_total_analysis(data_dic_key='Time')
    data_total_analysis(data_dic_key='Sub_time')
    data_total_analysis(data_dic_key='Send_data_time')
    data_total_analysis(data_dic_key='RSRP')


# 获取Pcorelog的数据分析
def get_SNR_RSRP():
    global data_dict, data_dict0
    rsrp = []
    lsnr = []
    date = []
    data_dict1 = {"CS FAIL": 0, "CS SUCCESS": 0}

    path = file_list2[0]
    pattern = r"(?=.*lsnr)(?=.*rsrp)(?=.*nrspower)"
    pattern1 = r"CS SUCCESS"
    pattern2 = r"CS FAIL"
    pattern3 = r"haldfeTransmit"
    pattern4 = r"END="
    pattern5 = r"Tx finish"
    i = 0
    with open(path, 'r') as f:
        line = f.readline()
        while line:
            line = f.readline()
            result = re.search(pattern, line)
            result1 = re.search(pattern1, line)
            result2 = re.search(pattern2, line)
            result3 = re.search(pattern3, line)
            result4 = re.search(pattern4, line)
            result5 = re.search(pattern5, line)
            if result1 != None:
                data_dict1['CS SUCCESS'] += 1
            if result2 != None:
                data_dict1['CS FAIL'] += 1
                # print(data_dict)
            if result != None:
                # print(line)
                try:
                    t = re.search(r'......................', line).group().split('[')[1]
                # print(t)
                # t = datetime.datetime.strptime(t, "%m-%d %H:%M:%S")  # 将字符串转换为时间格式

                    r = int(re.search('rsrp:.\d+', line).group().split(':')[1])
                    if r < -150:
                        print(line)
                except Exception as e:
                    print(line)
                    print(e.__str__())
                    r = 0

                try:
                    snr = int(re.search('lsnr:\s\d+', line).group().split(': ')[1])
                    # print(snr)
                except Exception as e:
                    print(e.__str__())
                    # print(line)
                    snr = 0
                if snr != 0:
                    # rsrp.append(r)
                    lsnrs = int(math.log(snr / 4096, 10) * 10)
                    if -150 < r < -50 and -10 < lsnrs < 30:
                        try:
                            date_obj = datetime.datetime.strptime(t, "%m-%d %H:%M:%S.%f")
                        except:
                            print(t)
                            print(date_obj)
                            print(line)
                        lsnr.append(lsnrs)
                        rsrp.append(r)
                        date.append(date_obj)
    CS_SUCESS_rate = format(float(data_dict1['CS SUCCESS']/(data_dict1['CS SUCCESS'] + data_dict1['CS FAIL']))*100, '.2f')
    data_dict1.update({"CS成功率": str(CS_SUCESS_rate) + '%'})

    data_dict1['CS SUCCESS'] = str(data_dict1['CS SUCCESS'])
    data_dict1['CS FAIL'] = str(data_dict1['CS FAIL'])

    data_dict0.update(data_dict1)
    fig, axes = plt.subplots()
    axes.plot(date, lsnr, label="SNR", color='#3cb44b', linewidth=0.5)
    axes.legend(loc='upper left', fontsize=8)
    axes1 = plt.twinx()
    axes1.plot(date, rsrp, label="RSRP", color='red', linewidth=0.5)
    axes1.legend(loc='upper right', fontsize=8)
    axes.set_yticks([-10, -2, 2, 4, 6, 10, 15, 20, 30])
    axes1.set_yticks([-50, -70, -80, -90, -100, -110, -140])
    plt.gcf().autofmt_xdate()
    plt.tight_layout()
    plt.savefig('SNR_RSRP.png')


# 将png转换成pdf
def png_to_pdf(png_path, pdf_path):
    # 创建一个 PDF 画布
    c = canvas.Canvas(pdf_path, pagesize=(460, 350))  # 可根据需要设置页面大小

    # 将 PNG 图片绘制到 PDF 画布上
    c.drawImage(png_path, 0, 0, width=460, height=350)  # 可根据需要设置图片的宽度和高度

    # 保存 PDF 文件
    c.save()


# 获取mylog里面的数据分析
def ping_result():
    global data_dict3
    ping_time = []
    file_mylog = [file for file in os.listdir('.') if 'my_log' in file and file.endswith('.txt')]
    pattern = r"LPING: .+,\d+,(\d+),\d+,\d+"
    # pattern = r"LPING: .+,\d+,(\d+),\d+"  # 电信认证版本
    path = file_mylog[0]
    print(path)
    with open(path, 'r', encoding='UTF-8') as f:
        try:
            line = f.readline()
            while line:
                line = f.readline()
                matches = re.findall(pattern, line)
                if matches:
                    ping_time.append(int(matches[0]))
        except Exception as e:
            print(e)
            print(line)

    # ping时延进行统计并生成饼图----------------------------------------------------------------------
    # data_range = [0, 1000, 5000, 10000, float('inf')]
    data_range = [0, 300, 600, 1000, 2000, 5000, 10000, float('inf')]
    pie_data1 = [sum(1 for num in ping_time if data_range[i] <= num < data_range[i + 1]) for i in
                 range(len(data_range) - 1)]
    # labels1 = ['0-1s', '0-5s', '5-10s', '10s+']
    labels1 = ['0-0.3s', '0.3-0.6s', '0.6-1s', '1-2s', '2-5s.', '5-10', '10s+']

    ax6 = plt1.add_subplot(336)
    # plt.subplot(336)
    colors1 = ['lightblue', 'lightgreen', 'lightyellow', 'lightcoral']
    ax6.pie(pie_data1, explode=[0.02, 0.02, 0.02, 0.02, 0.02, 0.02, 0.02], labels=labels1, labeldistance=1.1, autopct='%1.1f%%')
    ax6.set_title(f'ping时延分布({len(ping_time)})')

    data_dict3.update({'ping时延分布': None})
    for j in range(len(labels1)):
        data_rate = round((pie_data1[j]/len(ping_time))*100, 2)
        # data_rate = float(format(pie_data1[j]/len(ping_time), '.2f'))*100
        pie_data1[j] = str(pie_data1[j]) + "(" + str(data_rate) + '%)'
    for j in range(len(labels1)):
        data_dict3[labels1[j]] = pie_data1[j]


    data = pd.Series(ping_time)
    data = data[(data != -1) & (data != 0)]

    # 求平均值
    mean_value = str(round(data.mean(), 2))
    # 求最大值
    max_value = str(round(data.max(), 2))
    # 求最小值
    min_value = str(round(data.min(), 2))
    # 求方差
    var_value = str(round(data.var(), 2))
    # 写入字典
    data_dict0["每个包ping平均值"] = mean_value
    data_dict0["每个包ping最大值"] = max_value
    data_dict0["每个包ping最小值"] = min_value
    # data_dict0["每个包ping方差"] = var_value

    # plt.legend(labels1, loc='upper right', fontsize=3)
    # plt.savefig("ping时延分布.pdf")


# 获取myreport里面的数据分析，
def get_report_to_excel():
    global psm_mode, data_dict, data_dict_a_pcore, data_dict0
    # 2. 读取名字为“PSM模式测试”的sheet
    df = pd.read_excel(file_list[0], sheet_name=0)
    # 获取行和列数量
    num = str(df.shape[0] - 3)  # 获取行数量
    # print(num)
    str_1 = '第' + num + '次'
    # print(str_1)
    #
    with open(file_list3[0], 'r', encoding='utf-8') as f:
        line = f.readline()
        while line:
            line = f.readline()
            if str_1 in line:
                i = 0
                match_i = 10000000
                j = 0
                while line:
                    line = f.readline()
                    matches = re.findall(r'\{(.+?)\}', line)
                    for match in matches:
                        # 去除字符串的单引号并分割键值对
                        match_i = i
                        pairs = match.split(", ")
                        for input_string in pairs:
                            # 计算冒号的数量
                            colon_count = input_string.count(":")
                            # 如果冒号的数量超过1个，则删除第一个冒号
                            if colon_count > 1:
                                pair = input_string.replace(":", "", 1)
                            else:
                                pair = input_string
                            # print(pair)
                            key, value = pair.split(": ")
                            value = int(value)
                            if value > 0:
                                key = key.strip("'")
                                data_dict_a_pcore[key] = str(value)

                    if match_i < i:
                        j += 1
                        try:
                            if j <= 3:
                                line2 = line.split(' -')[1]
                                key, value = line2.split(":")
                                value = int(value)
                                data_dict[key] = str(value)

                        except Exception as e:
                            print(e)
                        if psm_mode == '0':
                            if '长时间不入网' in line:
                                line3 = line.split(' -')[1].split('  ')
                                # print(line3)
                                for l in line3:
                                    key, value = l.split("：")
                                    # value = int(value)
                                    if '长时间不入网' in key:
                                        data_dict0[key] = str(value)
                                    else:
                                        data_dict[key] = str(value)

                        if psm_mode == '1':
                            if 'PSM进入休眠失败' in line:
                                line3 = line.split(' -')[1].split('  ')
                                # print(line3)
                                for l in line3:
                                    key, value = l.split("：")
                                    # value = int(value)
                                    data_dict0[key] = str(value)
                        if '云平台注册失败' in line:
                            # print(line)
                            line4 = line.split(' -')[1].split('  ')
                            # print(line4)
                            for l in line4:
                                key, value = l.split("：")
                                if psm_mode == '1':
                                    data_dict0[key] = value
                                else:
                                    data_dict[key] = value

                        if '异常复位统计' in line:
                            # print(line)
                            line5 = line.split(' ', 2)[2].split('  ')
                            # print(line5)
                            # print(line5)
                            for l in line5:
                                key, value = l.split(":")
                                data_dict[key] = value

                        if 'ping总丢包数' in line:
                            # print(line)
                            line5 = line.split(' -')[1].split('  ')
                            for l in line5:
                                if "：" in l:
                                    key, value = l.split("：")
                                    if 'ping总丢包数' in key or 'LPING' in key:
                                        data_dict0[key] = value
                                    else:
                                        data_dict[key] = value
                                if ': ' in l:
                                    key, value = l.split(": ")
                                    if 'ping总丢包数' in key or 'LPING' in key:
                                        data_dict0[key] = value
                                    else:
                                        data_dict[key] = value

                    i = i + 1


if __name__ == "__main__":
    version = ''
    # excel表第1列统计数据
    data_dict0 = {}
    # excel表第2列统计数据
    data_dict = {}
    # excel表第3列数据,t
    data_dict3 = {}
    # 读取myreport中acore和pcore监控的数据
    data_dict_a_pcore = {}
    # 非PSM读取excel里面的数据，生成自动按，标题为key，值为value
    doc1 = {}
    # PSM模式读取excel里面的数据，生成自动按，标题为key，值为value
    dict2 = {}
    # 统计丢包数，生成新sheet2
    dict_sheet2 = {}
    # 创建一个 Excel 文件
    # ping_result()
    book = Workbook()
    writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')
    # writer = pd.ExcelWriter('output.xlsx', engine='openpyxl', workbook=book)
    writer.book = book
    worksheet = book.active
    writer.sheets['output'] = worksheet

    # 定义两个画布
    plt1 = plt.figure()


    if len(file_png) != 0:
        if '非PSM' in file_png[0]:
            psm_mode = '0'
        else:
            psm_mode = '1'
    else:
        psm_mode = input("输入’1‘PSM模式，其他非PSM模式：")
    # writer = writer.book.get_writer()

    # psm_mode = input("输入’1‘PSM模式，其他非PSM模式：")
    # 将.png图片转换成.pdf
    if len(file_png) == 0:
        print("未找到符合条件的.png文件。")
        # 创建一个.pdf文件
        c = canvas.Canvas('report.pdf', pagesize=(500, 380))  # 可根据需要设置页面大小
        # 保存 PDF 文件
        c.save()
    else:
        png_path = file_png[0]
        pdf_path = 'report.pdf'
        png_to_pdf(png_path, pdf_path)

    # 创建 PdfMerger 对象
    merger = PdfMerger()

    # 分析excel里面的数据
    if psm_mode == '1':
        psm_analys()
    else:
        # 获取当前目录中所有.csv文件名包含"Test"的文件
        file_list = [f for f in os.listdir('.') if f.endswith('.csv') and 'Test' in f]
        # 遍历所有文件
        for file_name in file_list:
            # 读取名为"测试数据"的sheet
            df = pd.read_excel(file_name, sheet_name='测试数据', header=None)
        version1 = df.iat[0, 0]
        version2 = df.iat[0, 1]
        data_dict[version1] = version2
        non_psm_analys()

    # 获取report.txt文件里面的报告，写入excel
    get_report_to_excel()

    # 创建 pandas 的 DataFrame 对象
    data_dict.update(data_dict_a_pcore)
    # print(data_dict0)

    # 调整第一组数据顺序字典data_dict
    ordered_data_dict = OrderedDict()
    for data_d in data_dict:
        ordered_data_dict[data_d] = data_dict[data_d]
    version0 = list(ordered_data_dict.keys())[0]
    ordered_data_dict.move_to_end('CFUN0_TIMEOUT', False)
    ordered_data_dict.move_to_end('OTA_RST', False)
    ordered_data_dict.move_to_end('PS_RST', False)
    ordered_data_dict.move_to_end('AT_RST', False)
    ordered_data_dict.move_to_end('WDG_RST', False)
    ordered_data_dict.move_to_end('HW_RST', False)
    ordered_data_dict.move_to_end('-异常复位统计', False)
    ordered_data_dict.move_to_end(version0, False)
    # 写入Excel的第一组数据
    df1 = pd.DataFrame(list(ordered_data_dict.items()), columns=['key2', 'value2'])
    # 写入sheet2的数据
    df2 = pd.DataFrame(dict_sheet2)

    # 写入 Excel 文件
    # df0.to_excel(writer, sheet_name='output', index=False, startcol=2, startrow=1)
    df1.to_excel(writer, sheet_name='output', index=False, startcol=0, startrow=0)
    df2.to_excel(writer, sheet_name='sheet2', index=False, startcol=0, startrow=0)
    # plt.show()
    # 保存 Excel 文件
    writer.save()

    # 添加第一个 PDF 文件
    merger.append('report.pdf')
    if psm_mode == '1':
        pdf_path = 'PSM_data_analist.pdf'
        png_to_pdf('PSM_data_analist.png', pdf_path)
        merger.append('PSM_data_analist.pdf')
    else:
        pdf_path = 'data_analist.pdf'
        png_to_pdf('data_analist.png', pdf_path)
        merger.append('data_analist.pdf')

    # 分析Pccorelog
    if len(file_list2) == 0:
        print("未找到Pcorelog文件。")
    else:
        get_SNR_RSRP()
        # 添加第二个 PDF 文件
        # pdf2_path = 'path/to/second.pdf'
        pdf_path = 'SNR_RSRP.pdf'
        png_to_pdf('SNR_RSRP.png', pdf_path)
        merger.append('SNR_RSRP.pdf')

    # 获取工作表和图片
    worksheet = writer.sheets['output']
    img1 = Image(file_png[0])

    # 调整图像大小
    img1.width = 540
    img1.height = 380

    img2 = Image('SNR_RSRP.png')

    if psm_mode == '1':
        img3 = Image('PSM_data_analist.png')
    else:
        img3 = Image('data_analist.png')
    img2.width = 540
    img2.height = 380
    img3.width = 540
    img3.height = 380
    # 在指定单元格插入图片
    worksheet.add_image(img1, 'A40')
    worksheet.add_image(img2, 'A60')
    worksheet.add_image(img3, 'A80')
    excel_A = 80
    for file in os.listdir('.'):
        if 'RSRP-SNR折线图' in file and file.endswith('.png'):
            excel_A += 20
            img_n = Image(file)
            img_n.width = 540
            img_n.height = 380
            worksheet.add_image(img_n, 'A' + str(excel_A))
            # print(file)
            # print(img_n)
            # print(excel_A)

    # new_df = pd.DataFrame([version])
    # new_df.to_excel('output.xlsx', )

    # 调整第二组数据顺序data_dict0
    ordered_data_dict0 = OrderedDict()
    for data_d in data_dict0:
        ordered_data_dict0[data_d] = data_dict0[data_d]

    if psm_mode == '1':
        ordered_data_dict0.move_to_end('CS FAIL', False)
        ordered_data_dict0.move_to_end('CS SUCCESS', False)
        ordered_data_dict0.move_to_end('CS成功率', False)
        ordered_data_dict0.move_to_end('发送数据失败', False)
        ordered_data_dict0.move_to_end('订阅失败', False)
        ordered_data_dict0.move_to_end('云平台注册失败', False)
        ordered_data_dict0.move_to_end('psm入网失败', False)
        ordered_data_dict0.move_to_end('PSM进入休眠失败', False)
    else:
        ordered_data_dict0.move_to_end('CS FAIL', False)
        ordered_data_dict0.move_to_end('CS SUCCESS', False)
        ordered_data_dict0.move_to_end('CS成功率', False)
        ordered_data_dict0.move_to_end('+LPING:2', False)
        ordered_data_dict0.move_to_end('+LPING:1', False)
        ordered_data_dict0.move_to_end('ping总丢包数', False)
        ordered_data_dict0.move_to_end('长时间不入网', False)
    # 写入excel的第二组数据
    df0 = pd.DataFrame(list(ordered_data_dict0.items()), columns=['key1', 'value1'])

    # df0 = pd.DataFrame(list(data_dict0.items()), columns=['key1', 'value1'])
    df0.to_excel(writer, sheet_name='output', index=False, startcol=2, startrow=1)

    df3 = pd.DataFrame(data_dict3.items(), columns=['分布', '次数(占比)'])
    # print(data_dict3)
    df3.to_excel(writer, sheet_name='output', index=False, startcol=4, startrow=1)

    writer.save()

    # 设置列宽和行高
    worksheet.column_dimensions['A'].width = 20  # 设置第一列的宽度为15
    worksheet.column_dimensions['b'].width = 16  # 设置第一列的宽度为15
    worksheet.column_dimensions['c'].width = 20  # 设置第一列的宽度为15
    worksheet.column_dimensions['d'].width = 16  # 设置第一列的宽度为15
    worksheet.column_dimensions['e'].width = 20  # 设置第一列的宽度为20
    worksheet.column_dimensions['f'].width = 25  # 设置第一列的宽度为25
    worksheet.row_dimensions[2].height = 60  # 设置第一行的高度为20

    # 设置边框样式
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    # 应用边框样式到所有单元格
    for row in worksheet.iter_rows(min_row=1, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.border = border

   # 保存 Excel 文件
    writer.save()
    # 将两个 PDF 文件合并
    merger.write('report.pdf')
    # 关闭 PdfMerger 对象
    merger.close()
    # print(cell_pingloss_dict)
    # print(data_dict0)
    print("分析完成")


